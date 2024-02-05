import tkinter as tk
from tkinter import ttk, StringVar, IntVar, messagebox, Toplevel
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import webbrowser, os

file = ('D:/US_project/COMPLETE/employee records system.xlsx')

if not os.path.exists(file): #if the file is not in the filepath, the file will be created automatically
    wb = Workbook()
    ws = wb.active
    wb.save(file)
else:
    wb = Workbook()
    wb = load_workbook(file)
    ws = wb.active




def refresh():
    tree.delete(*tree.get_children())
    # data = get_updated_data()
    for each_cell in range(2, ws.max_row + 1):
  
        if each_cell % 2 == 0:
            tree.insert(parent='', index='end',tags=('evenrow',) 
                        ,values=(ws['A' + str(each_cell)].value
                                 , ws['B' + str(each_cell)].value
                                 , ws['C' + str(each_cell)].value
                                 , ws['D' + str(each_cell)].value
                                 , ws['E' + str(each_cell)].value
                                 , ws['F' + str(each_cell)].value
                                 , ws['G' + str(each_cell)].value
                                 , ws['H' + str(each_cell)].value))
        else:
            tree.insert(parent='', index='end', tags=('oddrow',) 
                        ,values=(ws['A' + str(each_cell)].value
                                 , ws['B' + str(each_cell)].value
                                 , ws['C' + str(each_cell)].value
                                 , ws['D' + str(each_cell)].value
                                 , ws['E' + str(each_cell)].value
                                 , ws['F' + str(each_cell)].value
                                 , ws['G' + str(each_cell)].value
                                 , ws['H' + str(each_cell)].value))
            

def clear(): #clear all the entries

        ref_ent.delete(0, 'end')
        name_ent.delete(0, 'end')
        eml_ent.delete(0, 'end')
        select_gender.delete(0, 'end')
        pos.delete(0, 'end')
        con_ent.delete(0, 'end')
        sal_ent.delete(0, 'end')
        add_ent.delete(1.0, 'end')
        refresh()
        edit_btn.grid_forget()
        del_btn.grid_forget()
        save_btn.grid(row=9, column=0, padx=(5,0), pady=(0,5), sticky='e')
       
def save():
    id = ref_ent.get()
    name = name_ent.get()
    email = eml_ent.get()
    gender = select_gender.get()
    position = pos.get()
    contact = con_ent.get()
    salary = sal_ent.get()
    address = add_ent.get(1.0, 'end')

    Already_exist = False
    for every_row in range(2, ws.max_row + 1) :
        if id == ws['A' + str(every_row)].value:
            Already_exist =True
            break
       
    if Already_exist == True:
        messagebox.showerror("NOTICE", f"The id {id} is already exist")
        clear()
    
    elif (id == ""   #if the user don't type anything and pressed the save button an error will be triggered
          or name == "" 
          or email == "" 
          or gender == "" 
          or position == "" 
          or contact == "" 
          or salary == "" 
          or address == ""):
        messagebox.showerror("INVALID", "please complete your information")
    
    else:
        lr = str(ws.max_row + 1)

        ws['A'+lr] = id
        ws['B'+lr] = name
        ws['C'+lr] = email
        ws['D'+lr] = gender
        ws['E'+lr] = position
        ws['F'+lr] = contact
        ws['G'+lr] = salary
        ws['H'+lr] = address.title()
        
        messagebox.showinfo("NOTICE!","SAVED")
        
        refresh()#to get the updated data
    wb.save(file)#to save information typed by the user
        

def edit():
    # r.withdraw()
    for each_cell in range(2, (ws.max_row) + 1):
        if (ref_ent.get() == ws['A' + str(each_cell)].value):
            #found is a boolean variable that will trigger the stopping point of the loop
            Found = True 
            break 
        else:
            Found = False
    
    if(Found == True):
        # Top Level is python tkinter frame that can be invoke/called as part of another frame
        # Edit_form = Toplevel()
        # Edit_form.title("Edit Data From Excel")

        ef = tk.Frame(f)
        if ef.winfo_ismapped():
            ef.grid_forget()
        else:
            lf.grid_forget()
            fl.grid_forget()
            show_btn.grid_forget()
            treeFrame.grid_forget()
            owner.grid_forget()
            del_btn.grid_forget()
            ef.grid(row=1, column=0, padx=(5,10), pady=5)  # Show the frame

        global el
        el = tk.LabelFrame(ef,bg="khaki")
        el.grid(row=0, column=0)

        idExcel = StringVar()  
        nameExcel = StringVar()
        emailExcel = StringVar()
        genderExcel = StringVar()
        posExcel =  StringVar()
        contactExcel = StringVar()
        salaryExcel = StringVar()
        addressExcel = StringVar()

        #id
        global idlbl
        idlbl = tk.Label(el, width=7
                         , text="ID No.", bg="khaki", fg='#003f5c'
                         , font=('calibri bold', 12))
        idlbl.grid(row=0, column=0, padx=5, pady=5, sticky='e')

        idtxt = tk.Entry(el, width=30, textvariable=idExcel
                         , bg='LightCyan'
                         , font=('calibri bold', 10))
        idtxt.grid(row=0, column=1, padx=(5,10), pady=5, sticky='w')
        global idChk
        idChoice = IntVar()
        idChk = tk.Checkbutton(el, variable=idChoice, command=lambda:get_existing_id()
                               , text="same as before", bg="khaki", fg='#003f5c'
                               , font=('calibri bold', 10))
        idChk.grid(row=1, column=0, padx=(5,10), pady=5, columnspan=2)


        #name
        global namelbl
        namelbl = tk.Label(el, width=7
                           , text="Name", bg="khaki", fg='#003f5c'
                           , font=('calibri bold', 12))
        namelbl.grid(row=3, column=0, padx=5, pady=5, sticky='e')

        nametxt = tk.Entry(el, width=30, textvariable=nameExcel
                           , bg='LightCyan'
                           , font=('calibri bold', 10))
        nametxt.grid(row=3, column=1, padx=(5,10), pady=5, sticky='w')
        global namechk
        nameChoice = IntVar()
        namechk = tk.Checkbutton(el, variable=nameChoice, command=lambda:get_existing_name()
                                 , text="same as before", bg="khaki", fg='#003f5c'
                                 , font=('calibri bold', 10))
        namechk.grid(row=4, column=0, padx=5, pady=5, columnspan=2)



        #email'
        global emaillbl
        emaillbl = tk.Label(el, width=7
                            , text="Email", bg="khaki", fg='#003f5c'
                            , font=('calibri bold', 12))
        emaillbl.grid(row=6, column=0, padx=5, pady=5, sticky='e')
        
        emailtxt = tk.Entry(el, width=30, textvariable=emailExcel
                            , bg='LightCyan'
                            , font=('calibri bold', 10))
        emailtxt.grid(row=6, column=1, padx=(5,10), pady=5, sticky='w')
        global emailchk
        emailChoice = IntVar()
        emailchk = tk.Checkbutton(el, variable=emailChoice, command=lambda:get_existing_email()
                                  , text="same as before", bg="khaki", fg='#003f5c'
                                  , font=('calibri bold', 10))
        emailchk.grid(row=7, column=0, padx=5, pady=5, columnspan=2)


        #gender
        global genderlbl
        genderlbl = tk.Label(el, width=7
                             , text="Gender", bg="khaki", fg='#003f5c'
                             , font=('calibri bold', 12))
        genderlbl.grid(row=9, column=0, padx=5, pady=5, sticky='e')
        
        gendertxt = ttk.Combobox(el, width=27
                                 , textvariable=genderExcel, values=gend_list)
        gendertxt.grid(row=9, column=1, padx=(5,10), pady=5, sticky='w')
        global genderchk
        genderChoice = IntVar()
        genderchk = tk.Checkbutton(el, variable=genderChoice, command=lambda:get_existing_gender()
                                   , text="same as before", bg="khaki", fg='#003f5c'
                                   , font=('calibri bold', 10))
        genderchk.grid(row=10, column=0, padx=5, pady=5, columnspan=2)



        #position
        global poslbl
        poslbl = tk.Label(el, width=7
                          , text="Position", bg="khaki", fg='#003f5c'
                          , font=('calibri bold', 12))
        poslbl.grid(row=12, column=0, padx=5, pady=5, sticky='e')

        postxt = ttk.Combobox(el, width=27
                              , textvariable=posExcel, values=pos_list)
        postxt.grid(row=12, column=1, padx=(5,10), pady=5, sticky='w')
        global poschk
        posChoice = IntVar()
        poschk = tk.Checkbutton(el, variable=posChoice, command=lambda:get_existing_position()
                                , text="same as before",bg="khaki", fg='#003f5c'
                                , font=('calibri bold', 10))
        poschk.grid(row=13, column=0, padx=5, pady=5, columnspan=2)



        #contact
        global conlbl
        conlbl = tk.Label(el, width=7
                          , text="Contact",bg="khaki", fg='#003f5c'
                          , font=('calibri bold', 12))
        conlbl.grid(row=15, column=0, padx=5, pady=5, sticky='e')

        contxt = tk.Entry(el, width=30, textvariable=contactExcel
                          , bg='LightCyan'
                          , font=('calibri bold', 10))
        contxt.grid(row=15, column=1, padx=(5,10), pady=5, sticky='w')
        global conchk
        conChoice = IntVar()
        conchk = tk.Checkbutton(el, variable=conChoice, command=lambda:get_existing_contact()
                                , text="same as before", bg="khaki", fg='#003f5c'
                                , font=('calibri bold', 10))
        conchk.grid(row=16, column=0, padx=5, pady=5, columnspan=2)



        #salary
        global sallbl
        sallbl = tk.Label(el, width=7
                          , text="Salary", bg="khaki", fg='#003f5c'
                          , font=('calibri bold', 12))
        sallbl.grid(row=18, column=0, padx=5, pady=5, sticky='e')

        saltxt = tk.Entry(el, width=30, textvariable=salaryExcel
                          , bg='LightCyan'
                          , font=('calibri bold', 10))
        saltxt.grid(row=18, column=1, padx=(5,10), pady=5, sticky='w')
        global salchk
        salChoice = IntVar()
        salchk = tk.Checkbutton(el, variable=salChoice, command=lambda:get_existing_salary()
                                , text="same as before",bg="khaki", fg='#003f5c'
                                , font=('calibri bold', 10))
        salchk.grid(row=19, column=0, padx=5, pady=5, columnspan=2)



        #address
        global addlbl
        addlbl = tk.Label(el, width=7
                          , text="Address", bg="khaki", fg='#003f5c'
                          , font=('calibri bold', 12))
        addlbl.grid(row=21, column=0, padx=5, pady=5, sticky='e')

        addtxt = tk.Entry(el, width=30, textvariable=addressExcel
                          , bg='LightCyan'
                          , font=('calibri bold', 10))
        addtxt.grid(row=21, column=1, padx=(5,10), pady=5, sticky='w')
        global addchk
        addChoice = IntVar()
        addchk = tk.Checkbutton(el, variable=addChoice, command=lambda:get_existing_address()
                                , text="same as before", bg="khaki", fg='#003f5c'
                                , font=('calibri bold', 10))
        addchk.grid(row=22, column=0, padx=5, pady=5, columnspan=2)

        separator = ttk.Separator(el)
        separator.grid(row=23, column=0, padx=10, pady=(5,10), sticky='we', columnspan=2)

    
        # break

        def get_existing_id():
            if idChoice.get() == 1:
                idOld = ref_ent.get()
                idExcel.set(idOld)
            elif idChoice.get() == 0:
                idExcel.set("")
        
        def get_existing_name():
            if nameChoice.get() == 1:
                nameOld = name_ent.get()
                nameExcel.set(nameOld)
            elif nameChoice.get() == 0:
                nameExcel.set("")

        def get_existing_email():
            if emailChoice.get() == 1:
                emailOld = eml_ent.get()
                emailExcel.set(emailOld)
            elif emailChoice.get() == 0:
                emailExcel.set("")

        def get_existing_gender():
            if genderChoice.get() == 1:
                genderOld = select_gender.get()
                gendertxt.insert(0, select_gender.get())
                genderExcel.set(genderOld)
            elif genderChoice.get() == 0:
                genderExcel.set("")

        def get_existing_position():
            if posChoice.get() == 1:
                posOld = pos.get()
                posExcel.set(posOld)
            elif posChoice.get() == 0:
                posExcel.set("")

        def get_existing_contact():
            if conChoice.get() == 1:
                conOld = con_ent.get()
                contactExcel.set(conOld)
            elif conChoice.get() == 0:
                contactExcel.set("")
        
        def get_existing_salary():
            if salChoice.get() == 1:
                salOld = sal_ent.get()
                salaryExcel.set(salOld)
            elif salChoice.get() == 0:
                salaryExcel.set("")

        def get_existing_address():
            if addChoice.get() == 1:
                addOld = add_ent.get(1.0, 'end')
                addressExcel.set(addOld)
            elif addChoice.get() == 0:
                addressExcel.set("")

        def update():
                if (idtxt.get()== "" 
                    or nametxt.get() == "" 
                    or emailtxt.get() == "" 
                    or gendertxt.get() == "" 
                    or postxt.get() == "" 
                    or contxt.get()== "" 
                    or saltxt.get() == "" 
                    or addtxt.get() == ""):
                    messagebox.showerror("INVALID ACTION","please provide the information")
                else:
                    ws['A' + str(each_cell)].value = idtxt.get()
                    ws['B' + str(each_cell)].value = nametxt.get()
                    ws['C' + str(each_cell)].value = emailtxt.get()
                    ws['D' + str(each_cell)].value = gendertxt.get()
                    ws['E' + str(each_cell)].value = postxt.get()
                    ws['F' + str(each_cell)].value = contxt.get()
                    ws['G' + str(each_cell)].value = saltxt.get()
                    ws['H' + str(each_cell)].value = addtxt.get()


                    wb.save('employee records system.xlsx')#save the changes
                    messagebox.showinfo("updated", "DATA HAS BEEN UPDATED")
                    refresh()#to get the updated data
        editbtn = tk.Button(el, width=10
                             , text="Update", command=update)
        editbtn.grid(row=24, column=0, padx=5, pady=(0,10), columnspan=2, sticky='we')
        editbtn.bind("<Enter>", on_enter)
        editbtn.bind("<Leave>", on_leave)

        def back():
            # Edit_form.withdraw()
            # r.deiconify()
                ef.grid_forget()
                back_btn.grid_forget()
                lf.grid(row=1, column=0, padx=(10,5), pady=(5,10))
                fl.grid(row=0, column=0, padx=(4,23), pady=5, columnspan=2)
                show_btn.grid(row=0, column=1, padx=(5,10), pady=5, columnspan=3, sticky='e')
                # treeFrame.grid(row=1, column=1, padx=(5,10), pady=5)
                del_btn.grid(row=9, column=1,padx=(0,17.5), pady=(0,5), sticky='w')
                
        back_btn = tk.Button(f, width=7
                             , text="Back", command=back)
        back_btn.grid(row=2, column=0, padx=10, pady=10, sticky='e')
        back_btn.bind("<Enter>", on_enter)
        back_btn.bind("<Leave>", on_leave)

    else:
        messagebox.showerror("INVALID", "Please enter the ID No. if you wish to edit")#the user should enter the ID No. and search it before pressing the edit button
        r.deiconify()

def delete():

    found = False
    for every_row in range(2, ws.max_row + 1):
        if (ref_ent.get() == ws['A' + str(every_row)].value):
            found = True
            cell_found = every_row
            break
    if found == True:
        ws.delete_rows(cell_found)
        messagebox.showinfo("NOTICE", f"The data about employee {ref_ent.get()} is deleted!")#indication that the file was successfully deleted
        clear() #clear the treeview first
        refresh() #then refresh to get the updated data
    else:
        messagebox.showerror("ERROR", "Kindly type the existing ID No. if you want to delete")
        
    wb.save(file) #to save everything that changed within the workbook

def clear_tree():
    for item in tree.get_children():
        tree.delete(item)


def search():

        Found = False
        for cell in range(2, ws.max_row + 1):
            if (search_ent.get() == ws['A' + str(cell)].value):
                Found = True
                global cell_add
                cell_add = str(cell)
                break

            else:
                Found = False
           
        if (Found == True):
            clear()
            ref_ent.insert(0, ws['A' + cell_add].value)
            name_ent.insert(0, ws['B' + cell_add].value)
            eml_ent.insert(0, ws['C' + cell_add].value)
            select_gender.insert(0, ws['D' + cell_add].value)
            pos.insert(0, ws['E' + cell_add].value)
            con_ent.insert(0, ws['F' + cell_add].value)
            sal_ent.insert(0, ws['G' + cell_add].value)
            add_ent.insert(1.0, ws['H' + cell_add].value)
            save_btn.grid_forget()
            edit_btn.grid(row=9, column=0, padx=(5,0), pady=(0,5), sticky='e')
            del_btn.grid(row=9, column=1,padx=(0,17.5), pady=(0,5), sticky='w')
            # messagebox.showinfo("NOTICE", f"The data about {ref_ent.get()} exist in cell "+ cell_add)
            clear_tree()
            tree.tag_configure('search'
                               , background='Coral'
                               , font=('calibri', 8))
            tree.insert(parent='', index='end',tags=('search',) 
                        ,values=(ws['A' + str(cell)].value #ID No.
                                 , ws['B' + str(cell)].value #Name
                                 , ws['C' + str(cell)].value #Email
                                 , ws['D' + str(cell)].value #Gender
                                 , ws['E' + str(cell)].value #Position
                                 , ws['F' + str(cell)].value #Contact No.
                                 , ws['G' + str(cell)].value #Salary
                                 , ws['H' + str(cell)].value)) #Address
        
        else:
            messagebox.showerror("INVALID", f"Please enter valid ID No. to search")

def show():
    if treeFrame.winfo_ismapped(): #if the treeview is in unshow state the underneath buttons will be unshow to
        treeFrame.grid_forget()  # Hide the frame
        owner.grid_forget() #hide the dll secret button
        del_btn.grid_forget() #hide the delete button
    else:
        treeFrame.grid(row=1, column=1, padx=(5,10), pady=5)  # Show the frame
        
        owner.grid(row=2, column=0, columnspan=2)
        
        
        
#this is for buttonbind
def on_enter(event):
    event.widget.config(bg='#003f5c', fg="khaki")

def on_leave(event):
    event.widget.config(bg='SystemButtonFace', fg="black")

r = tk.Tk()
r.title("Employee Records System")
# icon = tk.PhotoImage(file='bilog.png')
# r.tk.call('wm', 'iconphoto', r._w, icon)

f = tk.Frame(r, bg='LightGrey')
f.pack(expand=True)


#search components
fl = tk.LabelFrame(f)
fl.grid(row=0, column=0
        , padx=(4,23), pady=5, columnspan=2)


search_ent = tk.Entry(fl, width=22
                      , bg='LightCyan', fg='Maroon'
                      , font=('calibri bold', 10))
search_ent.insert(0, " Search ID No.")
search_ent.bind("<FocusIn>", lambda e:search_ent.delete(0, 'end') )
search_ent.grid(row=0, column=0, padx=(10,5), pady=5)


search_btn = tk.Button(fl, width=4
                       ,text="⌕" , bg='silver'
                       , font=('calibri bold', 10)
                       ,  command=search)
search_btn.grid(row=0, column=1, padx=5, pady=5)
search_btn.bind("<Enter>", on_enter)
search_btn.bind("<Leave>", on_leave)


show_btn = tk.Button(f, width=7, relief="flat"
                     , text='s͟h͟o͟w͟', bg='LightGrey', fg='black'
                     , font=('calibri bold', 10)
                     , command=show)
show_btn.grid(row=0, column=1, padx=(5,15), pady=5, columnspan=3, sticky='e')


lf = tk.LabelFrame(f,bg='khaki')
lf.grid(row=1, column=0, padx=(10,5), pady=(5,10))
#employee reference id
ref_lbl = tk.Label(lf, width=10
                   , text="ID No.",bg="khaki", fg='#003f5c'
                   , font=('calibri bold', 11))
ref_lbl.grid(row=0, column=0, padx=5, pady=(10,5))

ref_ent = ttk.Entry(lf, width=20)
ref_ent.grid(row=0, column=1, padx=5, pady=(10,5))

#name
name_lbl = tk.Label(lf, width=10
                    , text="Name",bg="khaki", fg='#003f5c'
                    , font=('calibri bold', 11))
name_lbl.grid(row=1, column=0, padx=5, pady=5)

name_ent = ttk.Entry(lf, width=20)
name_ent.grid(row=1, column=1, padx=5, pady=5)

#email
eml_lbl = tk.Label(lf, width=10
                   , text="Email", bg="khaki", fg='#003f5c'
                   , font=('calibri bold', 11))
eml_lbl.grid(row=2, column=0, padx=5, pady=5)

eml_ent = ttk.Entry(lf, width=20)
eml_ent.grid(row=2, column=1, padx=5, pady=5)

#gender
gend_lbl = tk.Label(lf, width=10
                    , text="Gender",bg="khaki", fg='#003f5c'
                    , font=('calibri bold', 11))
gend_lbl.grid(row=3, column=0, padx=5, pady=5)

global gend_list
gend_list = ["Male"
             , "Female"]
select_gender = ttk.Combobox(lf, width=17
                             , values=gend_list)
select_gender.grid(row=3, column=1, padx=5, pady=5)
#designation
pos_lbl = tk.Label(lf, width=10
                   , text="Position",bg="khaki", fg='#003f5c'
                   , font=('calibri bold', 11))
pos_lbl.grid(row=4, column=0, padx=5, pady=5)

global pos_list
pos_list = ["HR"
            , "Manager"
            , "General Manager"
            , "Production Associate"
            , "Staff"
            , "Jr.Staff"]
pos = ttk.Combobox(lf, width=17
                   , values=pos_list)
pos.grid(row=4, column=1, padx=5, pady=5)

#contact no.
con_lbl = tk.Label(lf, width=10, text="Contact No."
                   ,bg="khaki", fg='#003f5c'
                   , font=('calibri bold', 11))
con_lbl.grid(row=5, column=0, padx=5, pady=5)

con_ent = ttk.Entry(lf, width=20)
con_ent.grid(row=5, column=1, padx=5, pady=5)

#Salary
sal_lbl = tk.Label(lf, width=10, text="Salary"
                   ,bg="khaki", fg='#003f5c'
                   , font=('calibri bold', 11))
sal_lbl.grid(row=6, column=0, padx=5, pady=5)

sal_ent = ttk.Entry(lf, width=20)
sal_ent.grid(row=6, column=1, padx=5, pady=5)

#address
add_lbl = tk.Label(lf, width=10, text="Address"
                   ,bg="khaki", fg='#003f5c'
                   , font=('calibri bold', 11))
add_lbl.grid(row=7, column=0, padx=5, pady=5)

add_ent = tk.Text(lf
                  , width=15, height=2
                  , bg='LightCyan')
add_ent.grid(row=7, column=1, padx=5, pady=5)

sep = ttk.Separator(lf)
sep.grid(row=8, column=0, columnspan=2, padx=10, pady=10, sticky='we')

save_btn = tk.Button(lf, width=7
                     , text="SAVE", command=save
                     , font=('calibri bold', 10) )
save_btn.grid(row=9, column=0, padx=(5,0), pady=(0,5), sticky='e')
save_btn.bind("<Enter>", on_enter)
save_btn.bind("<Leave>", on_leave)

edit_btn = tk.Button(lf, width=7
                     , text='EDIT', command=edit
                     , font=('calibri bold', 10))
edit_btn.bind("<Enter>", on_enter)
edit_btn.bind("<Leave>", on_leave)

clear_btn = tk.Button(lf, width=7
                      , text="CLEAR", command=clear
                      , font=('calibri bold', 10))
clear_btn.grid(row=9, column=1,padx=(0,17.5), pady=(0,5), sticky='e')
clear_btn.bind("<Enter>", on_enter)
clear_btn.bind("<Leave>", on_leave)

#treeview
treeFrame = ttk.Frame(f)
#the frame position is in show function


# treeScroll = ttk.Scrollbar(treeFrame)
# treeScroll.pack(side="right", fill="y")
v = ttk.Scrollbar(treeFrame)
v.grid(row=0, column=1, sticky='ns')

h = ttk.Scrollbar(treeFrame ,orient='horizontal')
h.grid(row=1, column=0, sticky='we')

def expand_treeview(event):
    tree.column("#0", width=event.width)
r.bind("<Configure>", expand_treeview)


global tree

tree = ttk.Treeview(treeFrame, selectmode='browse', wrap=None,
                     xscrollcommand=h.set,
                     yscrollcommand=v.set)
tree.grid(row=0, column=0)

v.configure(command=tree.yview)
h.configure(command=tree.xview)

tree["columns"] = ("1"
                   , "2"
                   , "3"
                   , "4"
                   , "5"
                   , "6"
                   , "7"
                   , "8")

tree['show'] = 'headings'

tree.column("1", width=100, anchor='c')
tree.column("2", width=130, anchor='c')
tree.column("3", width=170, anchor='c')
tree.column("4", width=100, anchor='c')
tree.column("5", width=110, anchor='c')
tree.column("6", width=110, anchor='c')
tree.column("7", width=100, anchor='c')
tree.column("8", width=160, anchor='c')

tree.heading("1", text="ID")
tree.heading("2", text="Name")
tree.heading("3", text="Email")
tree.heading("4", text="Gender")
tree.heading("5", text="Position")
tree.heading("6", text="Contact No.")
tree.heading("7", text="Salary")
tree.heading("8", text="Address")

# treeScroll.config(command=tree.yview)

style = ttk.Style()
style.theme_use("clam")
style.configure('Treeview', rowheight=31,
                fieldbackground='lavender')
style.configure("Treeview.Heading"
                , background="blue", foreground="white"
                , font=('calibri bold', 10))

dark_mode = tk.BooleanVar()
dark_mode.set(False)

tree.tag_configure('evenrow'
                   , background='lightcyan', font=('calibri', 8))
                        #this is to achieve the alternating colors in treeview
tree.tag_configure('oddrow'
                   , background='skyblue', font=('calibri', 8))

for each_cell in range(2, ws.max_row + 1): #use for loop to continously changing colors
                               #the odd and even row will be representing its colors         
    if each_cell % 2 == 0:  
        tree.insert(parent='', index='end',tags=('evenrow',)  
                    ,values=(ws['A' + str(each_cell)].value
                             , ws['B' + str(each_cell)].value
                             , ws['C' + str(each_cell)].value
                             , ws['D' + str(each_cell)].value
                             , ws['E' + str(each_cell)].value
                             , ws['F' + str(each_cell)].value
                             , ws['G' + str(each_cell)].value
                             , ws['H' + str(each_cell)].value))
    else:
        tree.insert(parent='', index='end', tags=('oddrow',) 
                    ,values=(ws['A' + str(each_cell)].value
                             , ws['B' + str(each_cell)].value
                             , ws['C' + str(each_cell)].value
                             , ws['D' + str(each_cell)].value
                             , ws['E' + str(each_cell)].value
                             , ws['F' + str(each_cell)].value
                             , ws['G' + str(each_cell)].value
                             , ws['H' + str(each_cell)].value))

def toggle_dark_mode():
    if dark_mode.get():

        # Set dark mode colors
        style.configure("Treeview.Heading", background="DarkSlateGray", foreground="white", font=('calibri bold', 10))
        style.configure('Treeview',fieldbackground='LightGrey')
        search_ent.config(bg='LightGrey')
        fl.config(bg='lavender')
        dark_mode_button.config(text='LightMode', bg='#424242',fg='#F5F6CE')
        r.configure(background='grey')
        f.config(bg='#424242')
        lf.config(bg='#0B2F3A')
        owner.config(bg='#424242', fg='#F5F6CE')
        show_btn.config(bg='#424242', fg='#F5F6CE')
        ref_lbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        name_lbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        eml_lbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        gend_lbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        pos_lbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        con_lbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        sal_lbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        add_lbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        tree.tag_configure('evenrow', background='DarkGray', font=('calibri', 9))
        tree.tag_configure('oddrow', background='Gray', font=('calibri', 9))

        #edit frame
        f.config(bg='#424242')
        el.config(bg='#0B2F3A')
        idlbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        idChk.config(bg='#0B2F3A', fg='LightSeaGreen')
        namelbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        namechk.config(bg='#0B2F3A', fg='LightSeaGreen')
        emaillbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        emailchk.config(bg='#0B2F3A', fg='LightSeaGreen')
        genderlbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        genderchk.config(bg='#0B2F3A', fg='LightSeaGreen')
        poslbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        poschk.config(bg='#0B2F3A', fg='LightSeaGreen')
        conlbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        conchk.config(bg='#0B2F3A', fg='LightSeaGreen')
        sallbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        salchk.config(bg='#0B2F3A', fg='LightSeaGreen')
        addlbl.config(bg='#0B2F3A', fg='LightSeaGreen')
        addchk.config(bg='#0B2F3A', fg='LightSeaGreen')
   
    else:
        # Set light mode colors
        style.configure("Treeview.Heading", background="blue", foreground="white", font=('calibri bold', 10))
        style.configure('Treeview',fieldbackground='lavender')
        search_ent.config(bg='LightCyan')
        fl.config(bg='Lavender')
        dark_mode_button.config(text='DarkMode', bg='lightgrey',fg='black')
        show_btn.config(bg='lightgrey', fg='black')
        owner.config(bg='lightgrey', fg='black')
        r.configure(background='Lavender')
        f.config(bg='LightGrey')
        lf.config(bg='khaki')
        ref_lbl.config(bg="khaki", fg='#003f5c')
        name_lbl.config(bg="khaki", fg='#003f5c')
        eml_lbl.config(bg="khaki", fg='#003f5c')
        gend_lbl.config(bg="khaki", fg='#003f5c')
        pos_lbl.config(bg="khaki", fg='#003f5c')
        con_lbl.config(bg="khaki", fg='#003f5c')
        sal_lbl.config(bg="khaki", fg='#003f5c')
        add_lbl.config(bg="khaki", fg='#003f5c')
        tree.tag_configure('evenrow', background='lightcyan', font=('calibri', 9))
        tree.tag_configure('oddrow', background='skyblue', font=('calibri', 9))

        #edit frame
        f.config(bg='LightGrey')
        el.config(bg='khaki')
        idlbl.config(bg="khaki", fg='#003f5c')
        idChk.config(bg="khaki", fg='#003f5c')
        namelbl.config(bg="khaki", fg='#003f5c')
        namechk.config(bg="khaki", fg='#003f5c')
        emaillbl.config(bg="khaki", fg='#003f5c')
        emailchk.config(bg="khaki", fg='#003f5c')
        genderlbl.config(bg="khaki", fg='#003f5c')
        genderchk.config(bg="khaki", fg='#003f5c')
        poslbl.config(bg="khaki", fg='#003f5c')
        poschk.config(bg="khaki", fg='#003f5c')
        conlbl.config(bg="khaki", fg='#003f5c')
        conchk.config(bg="khaki", fg='#003f5c')
        sallbl.config(bg="khaki", fg='#003f5c')
        salchk.config(bg="khaki", fg='#003f5c')
        addlbl.config(bg="khaki", fg='#003f5c')
        addchk.config(bg="khaki", fg='#003f5c')

dark_mode_button = tk.Checkbutton(f, text="Dark Mode"  #dark mode so that the program will be comfortable to use even at night
                                  , variable=dark_mode, command=toggle_dark_mode
                                  ,bg='lightgrey')
dark_mode_button.grid(row=2, column=0, padx=5, pady=5, sticky='w')


def fb():
    webbrowser.open_new("https://www.facebook.com/DLLOnlinePage?mibextid=ZbWKwL") #will be directed to DLL Facebook Page
    r.deiconify()

owner = tk.Button(f, width=15, text="studying here at D̲L̲L̲"
                  , font=('calibri bold', 7)
                  , command=fb, relief="flat"
                  , bg='lightgrey', fg='black')



del_btn = tk.Button(lf, width=7, text="DELETE", command=delete, font=('calibri bold', 10))
del_btn.bind("<Enter>", on_enter)
del_btn.bind("<Leave>", on_leave)




r.mainloop()